# arquivo = workbook
# sheet = planilha
# cell = celula

from openpyxl import Workbook, load_workbook
'''
# selecionando a planilha
wb = load_workbook('planilha_teste.xlsx') # caso o arquivo não esteja no mesmo diretorio é preciso colocar o endereço inteiro

# nota: não é possível salvar um arquivo xlsx se ele estiver aberto

# acessando a planilha 'ativa' do arquivo excel
ws = wb.active

print(ws['B2'].value) # o valor de B2

ws['A2'].value = "Test" #muda o valor da celular A2 para Test. Só tem efeito depois que salvar a planilha
ws['A2'] = 'Test' # tbm funciona

# salvando a planilha. aqui se vc colocar outro nome criará uma planilha nova.
wb.save('planilha_teste_alterada.xlsx')

# mostrar os nomes das planilhas dentro do arquivo
print(wb.sheetnames) 

# criando uma planilha no arquivo
wb.create_sheet('Teste') 
'''

# criando workbooks e os salvando
'''wb = Workbook()
ws = wb.active
ws.title = 'Data'

ws.append(['Tim', 'Is', 'Great', '!'])
wb.save('tim.xlsx')'''

# loop dentro da planilha

from openpyxl.utils import get_column_letter

wb = load_workbook('tim.xlsx')
ws = wb.active

'''for row in range(1, 11):
    for col in range(1, 5):
        char = get_column_letter(col)
        ws[char + str(row)] = char + str(row)

wb.save('tim.xlsx')'''

# merge and unmerge em cells

'''ws.merge_cells("A1:D1") #só fica o valor da 1ª cell
ws.unmerge_cells('A1:D1')
wb.save('tim.xlsx')
'''

# inserting and deleting rows

'''ws.insert_rows(7) # adiciona linha em branco na posição mencionada
ws.delete_rows(7)

ws.delete_cols(2) # é numerico e não alfabetico. as colunas chamam a,b,c.. etc

wb.save('tim.xlsx')'''

# copying and moving cells

ws.move_range('C1:D11', rows=2, cols=2)

wb.save('tim.xlsx')