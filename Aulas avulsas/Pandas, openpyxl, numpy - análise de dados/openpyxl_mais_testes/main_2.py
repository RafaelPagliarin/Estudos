from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font

# dict com os dados usados
data = {
    'Joe': {
        'math': 65,
        'science': 78,
        'english': 98,
        'gym': 89
    },
    'Bill': {
        'math': 55,
        'science': 72,
        'english': 87,
        'gym': 95
    },
    'Tim': {
        'math': 100,
        'science': 45,
        'english': 75,
        'gym': 92
    },
    'Sally': {
        'math': 30,
        'science': 25,
        'english': 45,
        'gym': 100
    },
    'Jane': {
        'math': 100,
        'science': 100,
        'english': 100,
        'gym': 60
    }
}

# criando o workbook e seu nome
wb = Workbook()
ws = wb.active
ws.title = 'Grades'

# criando e adicionando o 'cabeçalho'
headings = ['Name'] + list(data['Joe'].keys()) #escolheu um estudante aleatório para pegar as keys do dict
ws.append(headings)

# adicionando os dados do dict acima na spreadsheet
for person in data:
    grades = list(data[person].values())
    ws.append([person] + grades)

# adicionando a media das notas
for col in range(2, len(data['Joe']) + 2):
    char = get_column_letter(col)
    ws[char + '7'] = f'=SUM({char + "2"}:{char + "6"})/{len(data)}'

# mudar style das celulas.
# não existe maneira de mudar a linha/coluna inteira. é preciso alterar celula por celula
for col in range(1, 6):
    ws[get_column_letter(col) + '1'].font = Font(bold=True, color='000000FF')



# salvando o wb
wb.save('NewGrades.xlsx')
